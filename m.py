# importing the module
import pandas as pd
import openpyxl     
import os 

# reading the files
f1 = pd.read_excel("C:\\Users\\I527297\\Desktop\\m2\\fsn\\FSM - Design.xlsx")              #Address of design document
f2 = pd.read_excel("C:\\Users\\I527297\\Desktop\\m2\\fsn\\FSM - Execution.xlsx")           #Address of execution document

# merging the files
f3 = f2[["ID","Provider","Owner","Delegates","Control Design"]].merge(f1[["Control Design","Name","Description","Internal","Category","Organization","ISO22301_2019","ISO27001_2013","ISO27017_2015","ISO27018_2019"]],
								on = "Control Design", 
                                     how = "left")
									

# creating a new file
f3.to_excel("R.xlsx", index = False)

######################################################################################################################

df = pd.read_excel("C:\\Users\\I527297\\Desktop\\m2\\R.xlsx")     #Address of merged excel "R.xlsx"

col_name = 'Category'
unique_values = df[col_name].unique()

for unique_value in unique_values:
    fvalue = unique_value.replace('/', ' ')
    print(unique_value)
    # s = open('output\\'+fvalue+'.xlsx', 'w')
    df_output = df[df[col_name] == unique_value]
    output_path = os.path.join('output1\\', fvalue + '.xlsx')
    df_output.to_excel(output_path, sheet_name=fvalue, index=False)

########################################################################################################    

baseLocation = os.getcwd() #current working directory

SFSF_SOA_File_Path = os.path.join(baseLocation, "fsn\FSM-soa.xlsx") #File_Location
ExcelFindings = os.path.join(baseLocation, "output1") #folder location

def writeToExcel(LDLE_SDLE, S_LE, SDSE, sheetname, rowNumber):
    sheetname.cell(row=rowNumber, column=8).value = SDSE
    sheetname.cell(row=rowNumber, column=9).value = LDLE_SDLE
    sheetname.cell(row=rowNumber, column=10).value = S_LE

def findIdsInExcel(ID_Map):
    df = pd.read_excel(SFSF_SOA_File_Path)
    Mapping_to_SAP_Controlls = df['Mapping to SAP Controlls'].tolist()

    excelSheetsToWrite = openpyxl.load_workbook(SFSF_SOA_File_Path, read_only=False, keep_vba=True) #writer
    sheetname = excelSheetsToWrite['Report']
    
    counter = 2
    for controlls in Mapping_to_SAP_Controlls:
        LDLE_SDLE = S_LE = SDSE = ""          #initialize as empty
        controllsList = str(controlls).splitlines()   #list of controls 
        for controlIds in controllsList:              #control id in list
            for excelSheet in ID_Map.keys():          
                if controlIds in ID_Map[excelSheet]:
                    if excelSheet == "Local Design, Local Execution (LDLE).xlsx" or excelSheet == "Shared Design, Local Execution (SDLE).xlsx":
                        LDLE_SDLE += controlIds + "\n"
                    elif excelSheet == "Shared Design, Shared Local Execution (SDS LE).xlsx":
                        S_LE += controlIds + "\n"
                    else:
                        SDSE += controlIds + "\n"
                    break
        writeToExcel(LDLE_SDLE, S_LE, SDSE, sheetname, counter)
        counter += 1
    excelSheetsToWrite.save('FSM-soa.xlsx')


def extractDataFromSheets():
    '''Method to Extract IDS from Excel Sheets in ExcelFindings Dir'''
    ID_Map = dict() #empty dict for storing ids

    #Listing all the excel sheets in ExcelFindings Dir
    for excelSheets in os.listdir(ExcelFindings):
        excelSheets_Path = os.path.join(ExcelFindings, excelSheets) #getting location of each excel sheet
        df_excelSheet = pd.read_excel(excelSheets_Path) #reading data from excel
        #Extracting ID Column and storing as value with key as excel sheet name in the Dict ID_MAP
        ID_Map[excelSheets] = df_excelSheet['Control Design'].tolist()
    findIdsInExcel(ID_Map)

if __name__ == "__main__":
    extractDataFromSheets()





        
